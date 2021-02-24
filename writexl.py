# The function is a slightly modified version of the one presented here:
# https://stackoverflow.com/questions/60372238/updating-excel-sheet-with-pandas-without-overwriting-the-file
# Here, 'first_col' is an added feature that together with 'startrow' helps defining the exact cell in the Excel sheet where the upper left corner of the Pandas 
# dataframe should be written in.

import pandas as pd
def append_df_to_excel(filename:str, df:pd.core.frame.DataFrame , sheet_name:str, startrow:int, first_col:int,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    import pandas as pd #version ==1.1.5
    import numpy as np #version ==1.18.1
    from openpyxl import load_workbook #version ==3.0.5                
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn"t exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: "/path/to/file.xlsx")
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: "Sheet1")
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl")

    # Python 2.x: define [FileNotFoundError] exception if it doesn"t exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    if "index" not in to_excel_kwargs:
        to_excel_kwargs["index"] = False

    try:
        # try to open an existing workbook
        if "header" not in to_excel_kwargs:
            to_excel_kwargs["header"] = True
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
            to_excel_kwargs["header"] = False

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        to_excel_kwargs["header"] = True

    if startrow is None:
        startrow = 0

    # write out the new sheet
    #startrow and first col are the upper row and left column from which the dataframe will start.
    df.to_excel(writer, sheet_name, startrow=startrow,startcol=first_col-1, **to_excel_kwargs)

    # save the workbook
    writer.save()
