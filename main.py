# importing libraries
import pandas as pd #version ==1.1.5
import numpy as np #version ==1.18.1
import sqlalchemy #version ==1.3.21
from openpyxl import load_workbook #version ==3.0.5
from itertools import islice
import os
from win32com.client import Dispatch
import writexl #wrtite.py file in the directory

################################## Establishing a connection to the mysql database#####################################################################################
username='username'
password='password'
host='localhost'
port='port_number'
database_name='laboratory'
engine = sqlalchemy.create_engine('mysql+pymysql://{}:{}@{}:{}/{}'.format(username,password,host,port,database_name))

###############Excel sheets ('Reaction', 'Reagents', 'Users') to the database tables ('reactions','reagents','users' )#################################################
#Name and directory of the excel file
filename=os.path.join(os.getcwd() , 'data\lab.xlsx')

# Describing few features from three of the tables in the database and their corresponding sheet source in the excel file
# db_table_dic= {
#   'excel_sheet_name':
#               [corresponding database table name, 
#               max_col to be read from each excel sheet, 
#               column names]}

db_table_dic = {
    'Reactions': 
                ['reactions', 
                 8, 
                ('Reaction_id', 'Date', 'Temperature', 'Pressure', 'Yield', 'Synth_Protocol', 'User_id', 'Reagent_id')], 
    'Reagents': 
                ['reagents', 
                7, 
                ('Reagent_id', 'CAS', 'Location_ID', 'Purchase_Date', 'Open_Date', 'Supplier_id', 'Price_Eur')], 
    'Users': 
        ['users', 4, ('User_id', 'Start_Date', 'First_Name', 'Family_Name')]}

# Iterating over the 3 excel sheets ('Reactions', 'Reagents', 'Users') to collect the data 
# and feed it to the corresponding tables in the MySQL Workbench database
for sheet_name in ('Reactions', 'Reagents', 'Users'):
    
    db_table_name=db_table_dic[sheet_name][0]
    col_max=db_table_dic[sheet_name][1]
    cols=db_table_dic[sheet_name][2]
    
    # For every excel sheet, there is a corresponding table in the database.
    # Side note: There are more tables in the database, given that the data are normalized for many-to-many relationships among Reactions, Users and Reagents.
    # Here, the code reads the corresponding table and saves it in a dataframe called 'database'. 
    database = pd.read_sql_table(db_table_name, engine)
    
    #because of the dealing with excel, all indices are set to start from 1
    database.index = np.arange(1, len(database)+1)
    
    # openpyxl library accesses the worksheet
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    data = ws.values

    col_min=1; # col_max is defined in db_table_dic
    row_min=1; row_max=ws.max_row
    
    data = list(data)[row_min:row_max]

    # Because of the work with excel, all indices are set to start from 1
    idx = [r[0] for r in data]

    #       islice(iterable,   start,        stop)  #col_min-1 has -1 for the reason that counting starts from 1 but python starts from 0.
    data = (islice(r,          col_min-1,    col_max) for r in data)
            
    xlsx = pd.DataFrame(data, index=idx, columns=cols)
    
    if sheet_name=='Reactions':
        # MySQL database saves the temperature values in Kelvin
        xlsx['Temperature']=xlsx['Temperature']+273.15;
        xlsx['Synth_Protocol']=xlsx['Synth_Protocol'].str.upper().str.strip();
        
        # A data frame to extract the data from 'User_id','Reagent_id'
        # Dataframe 'extract' will have 'Reaction_id' as its 1st column, 'User_id' as the 2nd and 'Reagent_id' as its 3rd column.
        extract=xlsx.iloc[:,[0,6,7]] 
        
        #making sure only the new rows of the excel that are not in the database are extracted.
        extract=extract[extract.iloc[:,0]> database.shape[0]].astype(str) 
        
        # 'Reactions' will provide its first 6 columns to the database table 'reactions'. Remember, the rest went to the dataframe 'extract'.
        xlsx=xlsx.iloc[:,0:6]
        
    if sheet_name=='Reagents':
        # making strings upper case and chopping off blank space from its both ends
        for i in ('CAS','Location_ID'): xlsx[i]=xlsx[i].str.upper().str.strip()
        
    if sheet_name=='Users':
        # making strings upper case and chopping off blank space from its both ends
        for i in ('First_Name','Family_Name'): xlsx[i]=xlsx[i].str.upper().str.strip()
    
    # filling up/updating 'reactions, reagents, users' tables in the database
    xlsx_add=xlsx[xlsx.iloc[:,0]> database.shape[0]];
    xlsx_add.to_sql(con=engine, name=db_table_name, if_exists='append', index=False);
    
    # delete the data, preparing them for the next loop
    del(wb,data, xlsx_add,xlsx)

############### Excel sheets ('Reaction') has 2 columns 'User_id' and 'Reagent_id' that will be fed to to the database 
# tables 'reactions_users' and 'reactions_reagents' by an intermediate dataframe called 'extract' ####################################################################

# Preoparing the iterations over the 'extract' dataframe to fill up the 
# tables 'reactions_users' and 'reactions_reagents' in the mysql database.
# column_dic assigns column names to the table names
column_dic= {
    'db_column':[('Reaction_id', 'User_id'), ('Reaction_id', 'Reagent_id')],
    'db_table_name':['reactions_users', 'reactions_reagents']}

# Iteration along all the rows in the dataframe 'extract':
# filling up the first and second columns in the 'reaction_reagent' and 'reaction_user' tables of the MySQL database:
for counter in (1, 2): #First, we will fil up the the first column and then the second column using 'counter'
    a,b=[],[] #[a] will read the values from the 'Reaction_id', [b] will read the values from 'User_id' and 'Reagent_id'
    for i in range(extract.shape[0]):  # i represents the number of iteration along all the rows in the dataframe 'extract'
        for j in range( len ( extract.iloc[ i , counter] .split(","))): # j represents the iterations in each cell, depending on how many numbers are comma-separated
            # Extracting the first column in the row = 'reaction_id' (.iloc[i,0])
            a.append(extract.iloc[i,0])
            #extracting the second column which is either 'User_id' or 'Reagent_id'. 'counter' will take care of the 1st and 2nd columns of 'extract'
            b.append( extract.iloc [ i , counter ].split(",")[j])
    # Pouring [a] and [b] in a dataframe called df
    df=pd.DataFrame(list(zip(a, b)), columns =column_dic['db_column'][counter-1])
    
    # Writing df to the MySQL database
    df.to_sql(con=engine, name=column_dic['db_table_name'][counter-1], if_exists='append', index=False)
    
    # Deleting df, preparing it for the next loop
    del(df)

################### Making Use of Views (Defined in MySQL Workbench) to fill up Overview sheet in the exel file ######################################################
# reagent_use :
# Shows
# 1) how many people have used each chemical and
# 2) how many reactions make use of each chemical.
query1 = '''
 SELECT *
 FROM reagent_use 
'''
reagent_use = pd.read_sql_query(query1,engine)

# user_metrics:
# Shows
# 1) how many syntheses has each person conducted
# 2) and on which dates¶
query2 = '''
 SELECT *
 FROM user_metrics 
'''
user_metrics = pd.read_sql_query(query2,engine)

#reactions_overview:
# Shows
# 1) On average, how many chemicals are used in each synthesis reaction
# 2) On average, how many people work on each synthesis reaction.
# 3) What is the average reaction temperature in the lab?
# 4) What is the average yield of the syntheses in the lab?
query3 = '''
 SELECT *
 FROM reactions_overview 
'''
reactions_overview = pd.read_sql_query(query3,engine)

#Closing the excel file before saving 'overview' sheet in it is necessary.
xl = Dispatch('Excel.Application')
xl_file = xl.Workbooks.Open(filename)
xl_file.Close(True, filename)

# writing in the excel file's 'overview', using the writexl.py file in the directory
sheet_name='Overview'
df_dic={1:reagent_use, 6:user_metrics, 11:reactions_overview}
for i in df_dic.keys():
    writexl.append_df_to_excel(filename, df_dic[i], sheet_name, startrow=1,first_col=i, truncate_sheet=False)

 
